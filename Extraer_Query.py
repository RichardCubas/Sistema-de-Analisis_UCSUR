# Extraer_Query_GUI.py
# -*- coding: utf-8 -*-
"""
Extractor de Query UCSUR → Parquet + Excel con GUI (CustomTkinter)

FUNCIONALIDADES (NO SE ELIMINAN):
- Pregrado / CPE
- Cargar Query (CSV o Excel .xlsx) o continuar con Query existente
- Detección de encabezado (auto + manual)
- Normalización robusta y corrección ortográfica de cursos
- Unifica 'Matemática Básica'
- Guarda info del último query en info_query.txt:
    QUERY-PREGRADO, cargado el AAAA-MM-DD HH:MM
    QUERY-CPE, cargado el AAAA-MM-DD HH:MM
- Opción: Descargar Query filtrado actual (PARQUET)
- Opción: Generar y descargar datos filtrados (EXCEL, a demanda)

OPTIMIZACIÓN CLAVE (SOLUCIÓN DEL PROBLEMA PRINCIPAL):
- Para Excel grande (especialmente PREGRADO): NO convertir todo a CSV.
- Se hace Excel → Parquet temporal "recortado" (solo columnas necesarias + desde header),
  luego DuckDB filtra y genera parquet final.
- Si pyarrow no está, cae a Excel → CSV temporal (tu método anterior).
"""

import os
import re
import csv
import shutil
import tempfile
import unicodedata
from datetime import datetime

import duckdb
import pandas as pd

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk

# Excel streaming
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# Parquet streaming (OPCIONAL pero recomendado)
try:
    import pyarrow as pa
    import pyarrow.parquet as pq
except Exception:
    pa = None
    pq = None
# ============================================================
# MAPA CÓDIGO → CARRERA (UCSUR)
# ============================================================
MAPA_CARRERA = {
    "PEMSI": "INGENIERÍA EMPRESARIAL Y DE SISTEMAS",
    "PFADM": "ADMINISTRACIÓN DE EMPRESAS",
    "PFAGN": "AGRONOMÍA Y NEGOCIOS",
    "PFAHT": "ADMINISTRACIÓN HOTELERA Y TURISMO",
    "PFALI": "INGENIERÍA DE ALIMENTOS",
    "PFANI": "ADMINISTRACIÓN DE NEGOCIOS INTERNACIONALES",
    "PFAQI": "ARQUITECTURA DE INTERIORES",
    "PFAQU": "ARQUITECTURA Y URBANISMO AMBIENTAL",
    "PFARE": "ARTES ESCÉNICAS",
    "PFARS": "ADMINISTRACIÓN DE REDES Y SEGURIDAD INFORMÁTICA",
    "PFART": "ARTES ESCÉNICAS Y LITERATURA",
    "PFATS": "TURISMO SOSTENIBLE Y HOTELERÍA",
    "PFBMA": "BIOLOGÍA MARINA",
    "PFCIP": "CIENCIAS POLÍTICAS",
    "PFCIV": "INGENIERÍA CIVIL",
    "PFCMI": "COMUNICACIÓN AUDIOVISUAL",
    "PFCMK": "COMUNICACIÓN Y MARKETING",
    "PFCOC": "CONTABILIDAD CORPORATIVA",
    "PFCOE": "CONTABILIDAD EMPRESARIAL",
    "PFCOP": "COMUNICACIÓN Y PUBLICIDAD",
    "PFCYF": "CONTABILIDAD Y FINANZAS",
    "PFDER": "DERECHO",
    "PFDIS": "DISEÑO PROFESIONAL GRÁFICO",
    "PFENF": "ENFERMERÍA",
    "PFENI": "ECONOMÍA Y NEGOCIOS INTERNACIONALES",
    "PFEST": "ESTOMATOLOGÍA",
    "PFEYF": "ECONOMÍA Y FINANZAS",
    "PFFAR": "FARMACIA Y BIOQUÍMICA",
    "PFIAC": "INGENIERÍA ACUÍCOLA",
    "PFIAF": "INGENIERÍA AGROFORESTAL",
    "PFIAM": "INGENIERÍA AMBIENTAL",
    "PFIEN": "INGENIERÍA ECONÓMICA Y DE NEGOCIOS",
    "PFINA": "INGENIERÍA AGROFORESTAL",
    "PFIND": "INGENIERÍA INDUSTRIAL",
    "PFISE": "INGENIERÍA DE SISTEMAS EMPRESARIALES",
    "PFMEH": "MEDICINA HUMANA",
    "PFMKA": "MARKETING Y ADMINISTRACIÓN",
    "PFMKP": "MARKETING Y PUBLICIDAD",
    "PFMVZ": "MEDICINA VETERINARIA Y ZOOTECNIA",
    "PFNUT": "NUTRICIÓN Y DIETÉTICA",
    "PFOBS": "OBSTETRICIA",
    "PFPSI": "PSICOLOGÍA",
    "PFSOF": "INGENIERÍA DE SOFTWARE",
    "PFTRA": "TRADUCCIÓN E INTERPRETACIÓN PROFESIONAL",
    "PFBIM": "INGENIERÍA BIOMÉDICA",
    "PFIEC": "INGENIERÍA ELECTRÓNICA Y SE SOFTWARE",
    "PFMIN": "INGENIERÍA DE MINAS",
    "ADEOL": "ADMINISTRACIÓN DE EMPRESAS CPE ONLINE",
    "IEPAT": "INGENIERÍA ECONÓMICA Y DE NEGOCIOS PARA ADULTOS QUE TRABAJAN",
    "DEROL": "DERECHO CPE ONLINE",
    "NUPAP": "NUTRICIÓN Y DIETÉTICA PARA ADULTOS QUE TRABAJAN",
    "IESOL": "INGENIERÍA EMPRESARIAL Y DE SISTEMAS CPE ONLINE",
    "PSPAP": "PSICOLOGÍA PARA ADULTOS QUE TRABAJAN",
    "ISFOL": "INGENIERÍA DE SOFTWARE CPE ONLINE",
    "IENOL": "INGENIERÍA ECONÓMICA Y DE NEGOCIOS CPE ONLINE",
    "CVPAV": "INGENIERÍA CIVIL PARA ADULTOS QUE TRABAJAN",
    "EACYF": "CONTABILIDAD Y FINANZAS",
    "CYPOL": "COMUNICACIÓN Y PUBLICIDAD CPE ONLINE",
    "PSPAT": "PSICOLOGÍA PARA ADULTOS QUE TRABAJAN",
    "MKPAT": "MARKETING Y ADMINISTRACIÓN PARA ADULTOS QUE TRABAJAN",
    "NUPAT": "NUTRICIÓN Y DIETÉTICA PARA ADULTOS QUE TRABAJAN",
    "CPPAV": "COMUNICACIÓN Y PUBLICIDAD PARA ADULTOS QUE TRABAJAN",
    "MKAOL": "MARKETING Y ADMINISTRACIÓN CPE ONLINE",
    "DEPAT": "DERECHO PARA ADULTOS QUE TRABAJAN",
    "INDOL": "INGENIERÍA INDUSTRIAL CPE ONLINE",
}

# ============================================================
# CONFIG / CONSTANTES
# ============================================================

PARQUET_CENTRAL = "notas_filtradas_ucsur.parquet"
INFO_QUERY_TXT = "info_query.txt"

KEYS_HEADER = ["descr2", "short desc", "acad prog", "section", "id"]
MAX_SCAN_HEADER = 50  # filas a escanear para detectar encabezado
MAX_EXCEL_ROWS_SAFE_WARNING = 1_500_000  # aviso si excede esto (no bloquea)

# Columnas necesarias (según tu SELECT final)
NEEDED_COLS = [
    "ID",
    "LN,FN",
    "Descr2",
    "Section",
    "Acad Prog",
    "Responsable",
    "Short Desc",
    "Action Type",
]


# ============================================================
# NORMALIZACIÓN
# ============================================================
def normalizar_python(s):
    if not isinstance(s, str):
        s = str(s)
    s = s.lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def corregir_errores_curso(s):
    if not isinstance(s, str):
        return s

    s_norm = normalizar_python(s)

    reemplazos = {
        "fisica para ingenierios i":  "Física para Ingenieros I",
        "fisica para ingenierios ii": "Física para Ingenieros II",
        "fisica para ingenierios 1":  "Física para Ingenieros I",
        "fisica para ingenierios 2":  "Física para Ingenieros II",
        "fisica ing 1":               "Física para Ingenieros I",
        "fisica ing 2":               "Física para Ingenieros II",

        "matematica para ingenierios i":  "Matemática para Ingenieros I",
        "matematica para ingenierios ii": "Matemática para Ingenieros II",
        "matematica para ingenierios 1":  "Matemática para Ingenieros I",
        "matematica para ingenierios 2":  "Matemática para Ingenieros II",
        "Intro a las Ciencias Naturales": "Introducción a las Ciencias Naturales",
        "Intr. a las Ciencias Naturales": "Introducción a las Ciencias Naturales",
        "Introd a las Ciencias Naturale": "Introducción a las Ciencias Naturales",
        


        # Unificación explícita
        "matematica basica": "Matemática Básica",

        "ingenierios": "ingenieros",
    }

    return reemplazos.get(s_norm, s)


def cursos_pregrado():
    cursos = [
        "Álgebra", "Biología", "Bioquímica", "Estadística General",
        "Educación Ambiental", "Física", "Física I", "Física II",
        "Matemática", "Matemática General", "Matemática I", "Matemática II",
        "Matemática III", "Química", "Química General", "Química Orgánica"
    ]
    return [normalizar_python(c) for c in cursos]


def cursos_cpe():
    cursos = [
        "Estadística",
        "Introducción a las Ciencias Naturales",
        "Matemática Básica",
        "Matemática para Ingenieros I",
        "Matemática para Ingenieros II",
        "Física para Ingenieros I",
        "Física para Ingenieros II",
    ]
    return [normalizar_python(c) for c in cursos]


def obtener_norm_descr2_sql():
    # Normaliza Descr2 en DuckDB (sin tildes + trim + espacios + fix ingenierios->ingenieros)
    return """
    REGEXP_REPLACE(
        LOWER(
            REGEXP_REPLACE(
                TRIM(
                    REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                        REPLACE(REPLACE(REPLACE(REPLACE(REPLACE("Descr2",
                        'Á','A'),'É','E'),'Í','I'),'Ó','O'),'Ú','U'),
                        'á','a'),'é','e'),'í','i'),'ó','o'),'ú','u')
                ),
                '\\s+', ' ')
        ),
        'ingenierios',
        'ingenieros'
    )
    """


# ============================================================
# UTILIDADES GUI
# ============================================================
def append_log(log_widget, msg):
    try:
        if log_widget is not None:
            log_widget.configure(state="normal")
            log_widget.insert("end", msg + "\n")
            log_widget.see("end")
            log_widget.configure(state="disabled")
    except Exception:
        pass
    print(msg)


# ============================================================
# INFO DEL ÚLTIMO QUERY
# ============================================================
def guardar_info_query(modo):
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(INFO_QUERY_TXT, "w", encoding="utf-8") as f:
        f.write(f"QUERY-{modo.upper()}, cargado el {fecha}")


def leer_info_query():
    if not os.path.exists(INFO_QUERY_TXT):
        return None, None
    try:
        txt = open(INFO_QUERY_TXT, encoding="utf-8").read().strip()
        part1, part2 = txt.split(",", 1)
        modo = part1.replace("QUERY-", "").strip()
        fecha = part2.replace("cargado el", "").strip()
        return modo, fecha
    except Exception:
        return None, None


# ============================================================
# DETECTAR ENCABEZADO (CSV y XLSX)
# ============================================================
def _pedir_header_manual(parent):
    hdr_win = ctk.CTkToplevel(parent)
    hdr_win.title("Fila de encabezado")
    hdr_win.geometry("360x190")
    hdr_win.grab_set()

    label = ctk.CTkLabel(
        hdr_win,
        text="No se detectó la fila de encabezado.\n\n"
             "Ingrese el número de fila donde están los títulos\n(ejemplo: 7):"
    )
    label.pack(pady=10)

    entry = ctk.CTkEntry(hdr_win, width=120)
    entry.insert(0, "7")
    entry.pack(pady=5)

    result = {"value": None}

    def confirmar():
        try:
            fila = int(entry.get().strip())
            if fila <= 0:
                raise ValueError
            result["value"] = fila - 1  # a 0-index
            hdr_win.destroy()
        except ValueError:
            messagebox.showerror("Valor inválido", "Ingrese un entero positivo (ejemplo: 7).")

    btn_ok = ctk.CTkButton(hdr_win, text="Aceptar", command=confirmar)
    btn_ok.pack(pady=10)

    parent.wait_window(hdr_win)

    if result["value"] is None:
        raise ValueError("No se definió la fila de encabezado.")
    return result["value"]


def detectar_header_csv(file_path, parent, log_widget):
    append_log(log_widget, "🔍 Detectando fila de encabezado (CSV)...")

    header_row = None
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        for i in range(MAX_SCAN_HEADER):
            line = f.readline()
            if not line:
                break
            low = line.lower()
            if sum(k in low for k in KEYS_HEADER) >= 3:
                header_row = i
                break

    if header_row is None:
        append_log(log_widget, "⚠️ No se detectó automáticamente el encabezado (CSV).")
        header_row = _pedir_header_manual(parent)

    append_log(log_widget, f"✔ Encabezado detectado en fila {header_row + 1}")
    return header_row


def detectar_header_xlsx(xlsx_path, parent, log_widget):
    if load_workbook is None:
        raise RuntimeError("Falta openpyxl. Instale: pip install openpyxl")

    append_log(log_widget, "🔍 Detectando fila de encabezado (Excel)...")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    try:
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=MAX_SCAN_HEADER, values_only=True), start=1):
            joined = " ".join("" if v is None else str(v) for v in row).lower()
            if sum(k in joined for k in KEYS_HEADER) >= 3:
                header_row = i - 1  # 0-index
                break
    finally:
        wb.close()

    if header_row is None:
        append_log(log_widget, "⚠️ No se detectó automáticamente el encabezado (Excel).")
        header_row = _pedir_header_manual(parent)

    append_log(log_widget, f"✔ Encabezado detectado en fila {header_row + 1}")
    return header_row


# ============================================================
# UTILIDADES DuckDB path
# ============================================================
def _duckdb_path(p):
    return p.replace("\\", "/")


# ============================================================
# EXCEL → CSV TEMPORAL (FALLBACK)
# ============================================================
def excel_a_csv_temporal(xlsx_path, header_row_0idx, log_widget=None, progress_bar=None, parent=None):
    """
    Fallback: Convierte Excel → CSV temporal (streaming) DESDE header_row.
    (Mejor que antes porque ya no exporta filas basura antes del header)
    """
    if load_workbook is None:
        raise RuntimeError("Falta openpyxl. Instale: pip install openpyxl")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    try:
        max_row = ws.max_row or 0
        if max_row >= MAX_EXCEL_ROWS_SAFE_WARNING:
            append_log(log_widget, f"⚠️ Excel grande detectado: ~{max_row:,} filas. (CSV temporal puede tardar)")
    except Exception:
        pass

    fd, csv_path = tempfile.mkstemp(prefix="ucsur_query_", suffix=".csv")
    os.close(fd)

    append_log(log_widget, f"🧾 Excel → CSV temporal (desde header):\n   {csv_path}")

    try:
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            start_excel_row = header_row_0idx + 1  # openpyxl es 1-indexed
            # Escribimos desde la fila del header (incluida)
            for i, row in enumerate(ws.iter_rows(min_row=start_excel_row, values_only=True), start=1):
                writer.writerow(["" if v is None else v for v in row])

                if progress_bar and i % 5000 == 0:
                    try:
                        # Progreso aproximado (no perfecto)
                        if ws.max_row:
                            prog = min(0.95, (start_excel_row + i) / float(ws.max_row))
                            progress_bar.set(prog)
                        else:
                            progress_bar.set(0.2)
                        if parent:
                            parent.update_idletasks()
                    except Exception:
                        pass

    finally:
        wb.close()

    append_log(log_widget, "✔ CSV temporal generado.")
    return csv_path


# ============================================================
# EXCEL → PARQUET TEMPORAL RECORTADO (RÁPIDO)
# ============================================================
def excel_a_parquet_temporal_recortado(
    xlsx_path,
    header_row_0idx,
    log_widget=None,
    progress_bar=None,
    parent=None,
    sheet_index=0,
    batch_size=50000
):
    """
    Excel → Parquet temporal (streaming) SOLO con columnas necesarias (NEEDED_COLS)
    y SOLO desde la fila header.

    Requiere pyarrow. Si no está, debe usarse excel_a_csv_temporal().
    """
    if load_workbook is None:
        raise RuntimeError("Falta openpyxl. Instale: pip install openpyxl")
    if pq is None or pa is None:
        raise RuntimeError("Falta pyarrow. Instale: pip install pyarrow")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]

    fd, parquet_path = tempfile.mkstemp(prefix="ucsur_query_", suffix=".parquet")
    os.close(fd)

    append_log(log_widget, f"⚡ Excel → Parquet temporal RECORTADO:\n   {parquet_path}")
    append_log(log_widget, "   (solo columnas necesarias + desde header, para acelerar PREGRADO)")

    # localizar índices de columnas necesarias leyendo la fila del header
    header_excel_row = header_row_0idx + 1  # 1-indexed
    header_vals = None
    for row in ws.iter_rows(min_row=header_excel_row, max_row=header_excel_row, values_only=True):
        header_vals = ["" if v is None else str(v).strip() for v in row]
        break

    if not header_vals:
        wb.close()
        raise ValueError("No se pudo leer la fila de encabezado del Excel.")

    # Mapa nombre_columna -> índice
    name_to_idx = {}
    for idx, name in enumerate(header_vals):
        if name:
            name_to_idx[name] = idx

    missing = [c for c in NEEDED_COLS if c not in name_to_idx]
    if missing:
        wb.close()
        raise ValueError(
            "El Excel no contiene todas las columnas requeridas.\n"
            f"Faltan: {missing}\n\n"
            "Verifique que sea el query UCSUR estándar."
        )

    idxs = [name_to_idx[c] for c in NEEDED_COLS]

    writer = None
    total_written = 0

    # Buffers por columna
    buffers = {c: [] for c in NEEDED_COLS}

    try:
        max_row = ws.max_row or 0
        if max_row >= MAX_EXCEL_ROWS_SAFE_WARNING:
            append_log(log_widget, f"⚠️ Excel grande detectado: ~{max_row:,} filas. (pero parquet recortado es mucho más rápido)")

        data_start_row = header_excel_row + 1

        for r_i, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=1):
            # extraer solo las columnas necesarias
            for col_name, col_idx in zip(NEEDED_COLS, idxs):
                v = row[col_idx] if col_idx < len(row) else None
                buffers[col_name].append("" if v is None else v)

            if r_i % batch_size == 0:
                table = pa.table(buffers)
                if writer is None:
                    writer = pq.ParquetWriter(parquet_path, table.schema, compression="snappy")
                writer.write_table(table)
                total_written += batch_size
                buffers = {c: [] for c in NEEDED_COLS}

                if progress_bar:
                    try:
                        if max_row:
                            # aprox: filas procesadas / total
                            progress_bar.set(min(0.95, (data_start_row + r_i) / float(max_row)))
                        else:
                            progress_bar.set(0.3)
                        if parent:
                            parent.update_idletasks()
                    except Exception:
                        pass

        # flush final
        leftover = len(next(iter(buffers.values()))) if buffers else 0
        if leftover > 0:
            table = pa.table(buffers)
            if writer is None:
                writer = pq.ParquetWriter(parquet_path, table.schema, compression="snappy")
            writer.write_table(table)
            total_written += leftover

    finally:
        try:
            if writer is not None:
                writer.close()
        except Exception:
            pass
        wb.close()

    append_log(log_widget, f"✔ Parquet temporal recortado listo. Filas exportadas: {total_written:,}")
    return parquet_path


# ============================================================
# PROCESAR QUERY (CSV / XLSX) → PARQUET (FILTRADO)
# ============================================================
def procesar_query_archivo(file_path, parquet_especifico, modo,
                           parent, log_widget, progress_bar):
    """
    Problema principal resuelto:
    - Excel PREGRADO lento: ya NO se convierte todo a CSV.
      Ahora: Excel → Parquet temporal recortado → DuckDB filtra → Parquet final.
    - CSV sigue igual (DuckDB directo).
    """
    append_log(log_widget, f"\n📂 Archivo seleccionado:\n   {file_path}\n")

    if progress_bar:
        progress_bar.set(0)
        parent.update_idletasks()

    ext = os.path.splitext(file_path)[1].lower()

    temp_csv = None
    temp_parquet = None

    try:
        cursos_lista = cursos_pregrado() if modo == "pregrado" else cursos_cpe()
        curso_sql = ",".join([f"'{c}'" for c in cursos_lista])
        norm_descr2 = obtener_norm_descr2_sql()

        # -----------------------------------------
        # 1) CSV: DuckDB lee directo
        # -----------------------------------------
        if ext == ".csv":
            header_row = detectar_header_csv(file_path, parent, log_widget)
            append_log(log_widget, "📄 Fuente detectada: CSV (DuckDB directo)")

            fuente = f"read_csv_auto('{_duckdb_path(file_path)}', header=True, skip={header_row}, ignore_errors=true)"

        # -----------------------------------------
        # 2) Excel: Parquet temporal recortado (rápido)
        #    fallback a CSV temporal si no hay pyarrow
        # -----------------------------------------
        elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            if load_workbook is None:
                raise RuntimeError("Falta openpyxl. Instale: pip install openpyxl")

            header_row = detectar_header_xlsx(file_path, parent, log_widget)

            if pa is not None and pq is not None:
                append_log(log_widget, "📘 Fuente detectada: Excel")
                temp_parquet = excel_a_parquet_temporal_recortado(
                    xlsx_path=file_path,
                    header_row_0idx=header_row,
                    log_widget=log_widget,
                    progress_bar=progress_bar,
                    parent=parent,
                )
                fuente = f"read_parquet('{_duckdb_path(temp_parquet)}')"
            else:
                append_log(log_widget, "📘 Fuente detectada: Excel (sin pyarrow) → fallback CSV temporal")
                temp_csv = excel_a_csv_temporal(
                    xlsx_path=file_path,
                    header_row_0idx=header_row,
                    log_widget=log_widget,
                    progress_bar=progress_bar,
                    parent=parent,
                )
                # OJO: el CSV temporal ya empieza en el header, entonces skip=0
                fuente = f"read_csv_auto('{_duckdb_path(temp_csv)}', header=True, skip=0, ignore_errors=true)"

        else:
            raise ValueError("Formato no soportado. Use .csv o Excel (.xlsx/.xlsm).")

        append_log(log_widget, "⚙ Ejecutando consulta en DuckDB (filtro + reducción)…")

        query = f"""
        SELECT
            "ID"                  AS CodigoAlumno,
            "LN,FN"               AS Alumno,
            "Descr2"              AS Curso,
            "Section"             AS Seccion,
            "Acad Prog"           AS Carrera,
            "Responsable"         AS Docente,
            "Short Desc"          AS Evaluacion,
            "Action Type"::DOUBLE AS Nota
        FROM {fuente}
        WHERE {norm_descr2} IN ({curso_sql})
        """

        con = duckdb.connect()
        try:
            try:
                con.execute(f"PRAGMA threads={os.cpu_count() or 4}")
            except Exception:
                pass

            df = con.sql(query).df()

        finally:
            try:
                con.close()
            except Exception:
                pass

        append_log(log_widget, f"✔ Filas filtradas: {len(df):,}")

        if df.empty:
            messagebox.showwarning(
                "Sin resultados",
                "No se encontraron filas relevantes (cursos filtrados).\n"
                "Verifique que el archivo corresponda al formato UCSUR esperado."
            )
            return

        # Post-proceso ya sobre DF reducido
        df["Curso"] = (
            df["Curso"]
            .apply(corregir_errores_curso)
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df["Docente"] = df["Docente"].astype(str).str.strip().str.upper()
        # ------------------------------------------------------------
        # CARRERA: reemplazar código por nombre completo (si existe)
        # Si no existe en el mapa, se conserva el código original
        # ------------------------------------------------------------
        df["Carrera"] = (
            df["Carrera"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
        )

        df["Carrera"] = df["Carrera"].map(MAPA_CARRERA).fillna(df["Carrera"])

        df["Nota"] = pd.to_numeric(df["Nota"], errors="coerce").fillna(0)
        df = df.sort_values("Alumno")

        append_log(log_widget, f"📦 Guardando parquet específico:\n   {parquet_especifico}")
        df.to_parquet(parquet_especifico, index=False)

        append_log(log_widget, f"📦 Actualizando parquet central:\n   {PARQUET_CENTRAL}")
        df.to_parquet(PARQUET_CENTRAL, index=False)

        guardar_info_query(modo)
        append_log(log_widget, "💾 Información registrada en info_query.txt")

        if progress_bar:
            progress_bar.set(1.0)
            parent.update_idletasks()

        messagebox.showinfo(
            "Proceso completado",
            "El Query ha sido procesado correctamente.\n\n"
            "Ahora el sistema trabajará con el parquet optimizado.\n"
            "El Excel se genera solo bajo demanda."
        )

    finally:
        # limpieza temporales
        if temp_csv and os.path.exists(temp_csv):
            try:
                os.remove(temp_csv)
                append_log(log_widget, "🧹 CSV temporal eliminado.")
            except Exception:
                append_log(log_widget, "⚠️ No se pudo eliminar el CSV temporal (puede estar en uso).")

        if temp_parquet and os.path.exists(temp_parquet):
            try:
                os.remove(temp_parquet)
                append_log(log_widget, "🧹 Parquet temporal eliminado.")
            except Exception:
                append_log(log_widget, "⚠️ No se pudo eliminar el Parquet temporal (puede estar en uso).")


# ============================================================
# GENERAR EXCEL A PARTIR DEL PARQUET (RÁPIDO)
# ============================================================
def generar_excel_desde_parquet(parent, log_widget, progress_bar):
    central = PARQUET_CENTRAL

    if not os.path.exists(central):
        messagebox.showwarning(
            "Sin datos",
            "Aún no existe un Query filtrado.\nProcese un CSV/Excel o use un Query existente."
        )
        append_log(log_widget, f"⚠️ No se encontró {PARQUET_CENTRAL}")
        return

    append_log(log_widget, "📄 Leyendo datos del parquet central…")
    df = pd.read_parquet(central)

    modo, fecha = leer_info_query()
    if modo and fecha:
        ts = fecha.replace(":", "-").replace(" ", "_")
        default_name = f"Reporte_{modo.upper()}_{ts}.xlsx"
        append_log(log_widget, f"⬇ Último Query: QUERY-{modo.upper()}, cargado el {fecha}")
    else:
        default_name = "Reporte_filtrado.xlsx"
        append_log(log_widget, "⬇ Último Query: (sin información registrada)")

    output = filedialog.asksaveasfilename(
        title="Guardar Excel del Query filtrado actual",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel", "*.xlsx")]
    )

    if not output:
        append_log(log_widget, "❌ Descarga de Excel cancelada por el usuario.")
        return

    append_log(log_widget, "📊 Generando Excel institucional (modo rápido)…")

    if progress_bar:
        progress_bar.set(0)
        parent.update_idletasks()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Notas", index=False, header=False, startrow=4)

        wb = writer.book
        ws = writer.sheets["Notas"]

        if os.path.exists("logo_ucsur.png"):
            ws.insert_image("A1", "logo_ucsur.png", {"x_scale": 0.22, "y_scale": 0.22})

        ws.write(
            0, 2,
            "Reporte Académico – Cursos Básicos UCSUR",
            wb.add_format({"bold": True, "font_size": 14})
        )

        header_fmt = wb.add_format({
            "bold": True, "font_color": "white",
            "bg_color": "#003B70",
            "border": 1,
            "align": "center"
        })
        cell_fmt = wb.add_format({"border": 1})
        num_fmt = wb.add_format({"border": 1, "num_format": "0.00"})

        for col_idx, name in enumerate(df.columns):
            ws.write(3, col_idx, name, header_fmt)

        for col_idx, col_name in enumerate(df.columns):
            try:
                max_len = int(df[col_name].astype(str).str.len().max())
            except Exception:
                max_len = len(col_name)
            ancho = max(len(col_name), max_len) + 2
            fmt = num_fmt if col_name == "Nota" else cell_fmt
            ws.set_column(col_idx, col_idx, ancho, fmt)

    if progress_bar:
        progress_bar.set(1)
        parent.update_idletasks()

    append_log(log_widget, f"✅ Excel generado: {output}")
    messagebox.showinfo("Descarga completada", f"Excel guardado en:\n{output}")


# ============================================================
# GUI PRINCIPAL
# ============================================================
class ExtraerQueryApp(ctk.CTk):

    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("system")
        ctk.set_default_color_theme("blue")

        self.title("Extractor de Query UCSUR – Extraer_Query")
        self.geometry("800x600")
        self.minsize(800, 600)

        self.modo = None  # "pregrado" o "cpe"
        self.parquet_pre = "notas_filtradas_ucsur_pregrado.parquet"
        self.parquet_cpe = "notas_filtradas_ucsur_cpe.parquet"

        self._build_layout()
        self._actualizar_label_ultimo_query()

    def _build_layout(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header_frame = ctk.CTkFrame(self)
        header_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        header_frame.grid_columnconfigure(0, weight=1)
        header_frame.grid_columnconfigure(1, weight=1)

        title_label = ctk.CTkLabel(
            header_frame,
            text="Extractor de Query UCSUR",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)

        self.modo_label = ctk.CTkLabel(
            header_frame,
            text="Modo actual: (sin seleccionar)",
            font=ctk.CTkFont(size=14)
        )
        self.modo_label.grid(row=0, column=1, sticky="e", padx=10, pady=5)

        subtitulo = ctk.CTkLabel(
            header_frame,
            text="Pregrado / CPE • Query → Parquet + Excel (bajo demanda) • UCSUR 2025",
            font=ctk.CTkFont(size=13)
        )
        subtitulo.grid(row=1, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 2))

        self.lbl_ultimo_query = ctk.CTkLabel(
            header_frame,
            text="Último Query: (ninguno cargado aún)",
            font=ctk.CTkFont(size=13)
        )
        self.lbl_ultimo_query.grid(row=2, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 5))

        body_frame = ctk.CTkFrame(self)
        body_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        body_frame.grid_columnconfigure(0, weight=0)
        body_frame.grid_columnconfigure(1, weight=1)
        body_frame.grid_rowconfigure(0, weight=1)

        left_frame = ctk.CTkFrame(body_frame)
        left_frame.grid(row=0, column=0, sticky="ns", padx=10, pady=10)

        ctk.CTkLabel(left_frame, text="1. Seleccione el tipo de programa:", anchor="w").pack(
            anchor="w", padx=10, pady=(10, 5)
        )

        ctk.CTkButton(
            left_frame, text="PREGRADO", width=180,
            command=lambda: self.seleccionar_modo("pregrado")
        ).pack(padx=10, pady=5)

        ctk.CTkButton(
            left_frame, text="CPE", width=180,
            command=lambda: self.seleccionar_modo("cpe")
        ).pack(padx=10, pady=5)

        ctk.CTkLabel(left_frame, text="").pack(pady=5)

        ctk.CTkLabel(left_frame, text="2. Origen del Query:", anchor="w").pack(
            anchor="w", padx=10, pady=(10, 5)
        )

        self.btn_cargar = ctk.CTkButton(
            left_frame,
            text="Cargar CSV o Excel",
            width=180,
            command=self.cargar_query_archivo
        )
        self.btn_cargar.pack(padx=10, pady=5)

        self.btn_existente = ctk.CTkButton(
            left_frame,
            text="Usar Query existente",
            width=180,
            command=self.usar_query_existente
        )
        self.btn_existente.pack(padx=10, pady=5)

        ctk.CTkLabel(left_frame, text="").pack(pady=5)

        ctk.CTkLabel(
            left_frame,
            text="3. Descargar Query filtrado actual (PARQUET):",
            anchor="w"
        ).pack(anchor="w", padx=10, pady=(10, 5))

        self.btn_desc_parquet = ctk.CTkButton(
            left_frame,
            text="Descargar .parquet",
            width=180,
            command=self.descargar_query_actual
        )
        self.btn_desc_parquet.pack(padx=10, pady=5)

        ctk.CTkLabel(left_frame, text="").pack(pady=5)

        ctk.CTkLabel(
            left_frame,
            text="4. Descargar datos filtrados (EXCEL):",
            anchor="w"
        ).pack(anchor="w", padx=10, pady=(10, 5))

        self.btn_desc_excel = ctk.CTkButton(
            left_frame,
            text="Generar y descargar Excel",
            width=180,
            command=lambda: generar_excel_desde_parquet(self, self.txt_logs, self.progress)
        )
        self.btn_desc_excel.pack(padx=10, pady=5)

        ctk.CTkLabel(left_frame, text="").pack(pady=5)

        ctk.CTkButton(
            left_frame,
            text="Salir",
            fg_color="#B3261E",
            hover_color="#7F0000",
            width=180,
            command=self.destroy
        ).pack(padx=10, pady=(20, 10))

        right_frame = ctk.CTkFrame(body_frame)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        right_frame.grid_rowconfigure(1, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(right_frame, text="Registro de eventos:", anchor="w").grid(
            row=0, column=0, sticky="w", padx=10, pady=(10, 5)
        )

        self.txt_logs = ctk.CTkTextbox(right_frame, width=400, height=350)
        self.txt_logs.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        self.txt_logs.configure(state="disabled")

        self.progress = ctk.CTkProgressBar(right_frame, mode="determinate")
        self.progress.grid(row=2, column=0, sticky="ew", padx=10, pady=(10, 5))
        self.progress.set(0)

        self.status_label = ctk.CTkLabel(right_frame, text="Listo.", anchor="w")
        self.status_label.grid(row=3, column=0, sticky="w", padx=10, pady=(0, 10))

    def _actualizar_label_ultimo_query(self):
        modo, fecha = leer_info_query()
        if modo and fecha:
            self.lbl_ultimo_query.configure(
                text=f"Último Query: QUERY-{modo.upper()}, cargado el {fecha}"
            )
        else:
            self.lbl_ultimo_query.configure(text="Último Query: (ninguno cargado aún)")

    def seleccionar_modo(self, modo):
        self.modo = modo
        txt = "Modo actual: PREGRADO" if modo == "pregrado" else "Modo actual: CPE"
        self.modo_label.configure(text=txt)
        append_log(self.txt_logs, f"▶ Programa seleccionado: {txt.split(': ')[1]}")
        self.status_label.configure(text="Modo actualizado.")

    def cargar_query_archivo(self):
        if self.modo is None:
            messagebox.showwarning(
                "Modo no seleccionado",
                "Primero seleccione el tipo de programa: PREGRADO o CPE."
            )
            return

        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo de Query UCSUR",
            filetypes=[
                ("CSV / Excel", "*.csv *.xlsx *.xlsm *.xltx *.xltm"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx *.xlsm *.xltx *.xltm"),
            ]
        )

        if not file_path:
            append_log(self.txt_logs, "❌ No se seleccionó archivo.")
            self.status_label.configure(text="No se seleccionó archivo.")
            return

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            append_log(self.txt_logs, "\n📌 CSV detectado: DuckDB filtra directo.")
        else:
            append_log(self.txt_logs, "\n📌 Excel detectado:")
            append_log(self.txt_logs, "   • Se leerá en streaming")
            if pa is not None and pq is not None:
                append_log(self.txt_logs, "   • Excel → Parquet temporal recortado (rápido)")
            else:
                append_log(self.txt_logs, "   • (sin pyarrow) fallback: Excel → CSV temporal (más lento)")
            if load_workbook is None:
                messagebox.showerror(
                    "Falta dependencia",
                    "Para leer Excel necesita openpyxl.\n\nInstale:\n   pip install openpyxl"
                )
                return

        parquet_especifico = self.parquet_pre if self.modo == "pregrado" else self.parquet_cpe

        try:
            self.status_label.configure(text="Procesando query...")
            self.progress.set(0)
            self.update_idletasks()

            procesar_query_archivo(
                file_path=file_path,
                parquet_especifico=parquet_especifico,
                modo=self.modo,
                parent=self,
                log_widget=self.txt_logs,
                progress_bar=self.progress
            )

            self.status_label.configure(text="Proceso completado.")
            self._actualizar_label_ultimo_query()

        except Exception as e:
            append_log(self.txt_logs, f"❌ Error durante el procesamiento: {e}")
            self.status_label.configure(text="Error en el proceso.")
            messagebox.showerror("Error", f"Ocurrió un error:\n\n{e}")

    def usar_query_existente(self):
        if self.modo is None:
            messagebox.showwarning(
                "Modo no seleccionado",
                "Primero seleccione el tipo de programa: PREGRADO o CPE."
            )
            return

        parquet_especifico = self.parquet_pre if self.modo == "pregrado" else self.parquet_cpe

        if not os.path.exists(parquet_especifico):
            append_log(self.txt_logs, f"\n⚠️ No existe un Query previo para: {self.modo.upper()}")
            messagebox.showwarning(
                "Query inexistente",
                f"No se encontró:\n{parquet_especifico}\n\nPrimero debe generar uno cargando un CSV/Excel."
            )
            self.status_label.configure(text="Query inexistente.")
            return

        shutil.copyfile(parquet_especifico, PARQUET_CENTRAL)
        guardar_info_query(self.modo)

        append_log(self.txt_logs, f"\n📦 Usando Query existente ({parquet_especifico})")
        append_log(self.txt_logs, f"📦 {PARQUET_CENTRAL} actualizado")
        self.status_label.configure(text="Query existente cargado.")
        self._actualizar_label_ultimo_query()

        messagebox.showinfo(
            "Query existente",
            f"Se ha copiado el Query existente:\n{parquet_especifico}\n\na {PARQUET_CENTRAL}."
        )

    def descargar_query_actual(self):
        central = PARQUET_CENTRAL

        if not os.path.exists(central):
            messagebox.showwarning(
                "Sin Query filtrado",
                f"No se encontró '{PARQUET_CENTRAL}'.\n\nPrimero cargue o use un Query."
            )
            append_log(self.txt_logs, "⚠️ No hay Query filtrado actual para descargar (parquet).")
            return

        modo, fecha = leer_info_query()
        if modo and fecha:
            ts = fecha.replace(":", "-").replace(" ", "_")
            default_name = f"Query_{modo.upper()}_{ts}.parquet"
            append_log(self.txt_logs, f"⬇ Último Query registrado: QUERY-{modo.upper()}, cargado el {fecha}")
        else:
            default_name = PARQUET_CENTRAL
            append_log(self.txt_logs, "⬇ Último Query: (sin información registrada en info_query.txt)")

        destino = filedialog.asksaveasfilename(
            title="Guardar Query filtrado actual (.parquet)",
            defaultextension=".parquet",
            initialfile=default_name,
            filetypes=[("Parquet", "*.parquet")]
        )

        if not destino:
            append_log(self.txt_logs, "❌ Descarga de PARQUET cancelada por el usuario.")
            return

        try:
            shutil.copyfile(central, destino)
            append_log(self.txt_logs, f"📥 Parquet copiado correctamente a:\n{destino}")
            messagebox.showinfo("Descarga completada", f"Parquet guardado en:\n{destino}")
        except Exception as e:
            append_log(self.txt_logs, f"❌ Error al copiar PARQUET: {e}")
            messagebox.showerror("Error", f"No se pudo copiar el archivo:\n\n{e}")


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    app = ExtraerQueryApp()
    app.mainloop()
